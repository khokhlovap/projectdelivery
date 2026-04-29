import json
from django.shortcuts import render, redirect
from django.contrib.auth import login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse  
from django.utils import timezone  
from .forms import RegistrationForm
from django.contrib.auth.forms import AuthenticationForm
from delivery.models import Order, Client, Courier, OrderRating, OrderStatusHistory, User, CourierNotification, AIChatLog, Campaign 
from django.db.models import Count, Avg, Q
from datetime import datetime, timedelta
from django.core.paginator import Paginator
from django.contrib.auth.hashers import make_password
import random
import string
from django.core.mail import send_mail
from django.shortcuts import get_object_or_404
import re
import requests
import openpyxl
import os
from django.conf import settings
from django.http import HttpResponse
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer,
    Table, TableStyle
)
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from itertools import chain
from operator import attrgetter

# ЛК Клиент
def register_view(request):
    if request.user.is_authenticated:
        return redirect('accounts:home')
    
    if request.method == 'POST':
        form = RegistrationForm(request.POST)
        if form.is_valid():
            user = form.save()
            login(request, user)
            messages.success(request, f'Добро пожаловать, {user.get_full_name()}! Пожалуйста, заполните данные компании.')
            return redirect('accounts:company_setup')
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f'{error}')
    else:
        form = RegistrationForm()
    
    return render(request, 'registration/register.html', {'form': form})

def login_view(request):
    if request.user.is_authenticated:
        return redirect('accounts:home')
    
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            
            # Проверка роли пользователя и перенаправление
            if user.role == 'manager':
                return redirect('accounts:manager_dashboard')
            elif user.role == 'client':
                try:
                    if user.client_profile:
                        messages.success(request, f'Добро пожаловать, {user.get_full_name()}!')
                        return redirect('accounts:home')
                except:
                    messages.warning(request, 'Пожалуйста, заполните данные компании')
                    return redirect('accounts:company_setup')
            elif user.role == 'courier':
                # Если добавите панель курьера
                return redirect('delivery:courier_dashboard')
        else:
            messages.error(request, 'Неверный email или пароль')
    else:
        form = AuthenticationForm()
    
    return render(request, 'registration/login.html', {'form': form})

def logout_view(request):
    "Выход из системы"
    logout(request)
    return redirect('accounts:login')

@login_required
def company_setup(request):
    "Заполнение данных компании"
    try:
        client = request.user.client_profile
        return redirect('accounts:home')
    except:
        client = None
    
    if request.method == 'POST':
        Client.objects.create(
            user=request.user,
            company_name=request.POST.get('company_name'),
            inn=request.POST.get('inn'),
            kpp=request.POST.get('kpp'),
            legal_address=request.POST.get('legal_address'),
            actual_address=request.POST.get('actual_address'),
            company_phone=request.POST.get('company_phone'),
            company_email=request.POST.get('company_email'),
            bank=request.POST.get('bank'),
            settlement_account=request.POST.get('settlement_account'),
            correspondent_account=request.POST.get('correspondent_account'),
            contact_person_first_name=request.user.first_name,
            contact_person_last_name=request.user.last_name,
            contact_person_phone=request.user.phone,
            contact_person_email=request.user.email,
        )
        messages.success(request, 'Данные компании успешно сохранены! Теперь вы можете создавать заказы.')
        return redirect('accounts:home')
    
    return render(request, 'accounts/company_setup.html')

@login_required
def client_dashboard(request):
    "Главная страница личного кабинета"
    try:
        client_profile = request.user.client_profile
    except:
        messages.warning(request, 'Для создания заказов необходимо заполнить данные компании')
        return redirect('accounts:company_setup')
    
    try:
        active_orders = Order.objects.filter(
            client=client_profile,
            status__in=['created', 'pending', 'assigned', 'in_progress']
        ).order_by('-created_at')
        active_orders_count = active_orders.count()
        
        total_orders = Order.objects.filter(client=client_profile).count()
        delivered_count = Order.objects.filter(client=client_profile, status='delivered').count()
    except:
        active_orders = []
        active_orders_count = 0
        total_orders = 0
        delivered_count = 0
    
    context = {
        'active_orders': active_orders,
        'active_orders_count': active_orders_count,
        'total_orders': total_orders,
        'delivered_count': delivered_count,
        'active_tab': 'home',
    }
    return render(request, 'accounts/dashboard.html', context)

@login_required
def client_orders(request):
    "Все заказы клиента"
    try:
        client_profile = request.user.client_profile
        orders = Order.objects.filter(client=client_profile).order_by('-created_at')
        
        active_count = orders.filter(status__in=['created', 'pending', 'assigned', 'in_progress']).count()
        delivered_count = orders.filter(status='delivered').count()
        total_orders = orders.count()
        
        for order in orders:
            try:
                order.has_rating = hasattr(order, 'rating')
            except:
                order.has_rating = False
        
        paginator = Paginator(orders, 4)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)
        
    except:
        page_obj = []
        active_count = 0
        delivered_count = 0
        total_orders = 0
    
    context = {
        'page_obj': page_obj,
        'active_count': active_count,
        'delivered_count': delivered_count,
        'total_orders': total_orders,
        'total_count': orders.count() if 'orders' in locals() else 0,
        'active_tab': 'orders',
    }
    return render(request, 'accounts/clients_orders.html', context)

@login_required
def rate_order(request):
    "Оценка заказа клиентом"
    if request.method != 'POST':
        return JsonResponse({'error': 'Метод не разрешен'}, status=405)
    
    try:
        data = json.loads(request.body)
        order_id = data.get('order_id')
        rating = data.get('rating')
        
        if not rating or rating < 1 or rating > 5:
            return JsonResponse({'error': 'Некорректная оценка'}, status=400)
        
        order = Order.objects.get(id=order_id, client=request.user.client_profile)
        
        # Проверка, есть ли уже оценка
        if hasattr(order, 'rating'):
            return JsonResponse({'error': 'Оценка уже была оставлена'}, status=400)
        
        # Проверка, что заказ доставлен
        if order.status != 'delivered':
            return JsonResponse({'error': 'Оценку можно оставить только после доставки'}, status=400)
        
        # Создаем оценку
        OrderRating.objects.create(
            order=order,
            client=request.user.client_profile,
            courier=order.courier,
            rating=rating
        )
        
        return JsonResponse({'success': True, 'message': 'Спасибо за оценку!'})
    except Order.DoesNotExist:
        return JsonResponse({'error': 'Заказ не найден'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
@login_required
def check_order_rating(request):
    "Проверка, есть ли уже оценка у заказа"
    order_id = request.GET.get('order_id')
    
    try:
        order = Order.objects.get(id=order_id, client=request.user.client_profile)
        has_rating = hasattr(order, 'rating')
        return JsonResponse({'has_rating': has_rating})
    except Order.DoesNotExist:
        return JsonResponse({'has_rating': True, 'error': 'Заказ не найден'}, status=404)
    
@login_required
def client_settings(request):
    "Настройки профиля клиента"
    try:
        client = request.user.client_profile
    except:
        client = None
    
    if request.method == 'POST':
        action = request.POST.get('action')
        
        # Обновление личных данных
        if action == 'update_profile':
            user = request.user
            user.first_name = request.POST.get('first_name', '').strip()
            user.last_name = request.POST.get('last_name', '').strip()
            user.patronymic = request.POST.get('patronymic', '').strip()
            user.phone = request.POST.get('phone', '').strip()
            user.save()
            messages.success(request, 'Личные данные успешно обновлены')
            return redirect('accounts:client_settings')
        
        # Обновление данных компании
        elif action == 'update_company':
            if client:
                client.company_name = request.POST.get('company_name', '').strip()
                client.inn = request.POST.get('inn', '').strip()
                client.kpp = request.POST.get('kpp', '').strip()
                client.legal_address = request.POST.get('legal_address', '').strip()
                client.actual_address = request.POST.get('actual_address', '').strip()
                client.company_phone = request.POST.get('company_phone', '').strip()
                client.company_email = request.POST.get('company_email', '').strip()
                client.bank = request.POST.get('bank', '').strip()
                client.settlement_account = request.POST.get('settlement_account', '').strip()
                client.correspondent_account = request.POST.get('correspondent_account', '').strip()
                client.contact_person_first_name = request.POST.get('contact_person_first_name', '').strip()
                client.contact_person_last_name = request.POST.get('contact_person_last_name', '').strip()
                client.contact_person_patronymic = request.POST.get('contact_person_patronymic', '').strip()
                client.contact_person_phone = request.POST.get('contact_person_phone', '').strip()
                client.contact_person_email = request.POST.get('contact_person_email', '').strip()
                client.save()
                messages.success(request, 'Данные компании успешно обновлены')
            else:
                messages.error(request, 'Профиль компании не найден')
            return redirect('accounts:client_settings')
        
        # Смена пароля
        elif action == 'change_password':
            current_password = request.POST.get('current_password')
            new_password = request.POST.get('new_password')
            confirm_password = request.POST.get('confirm_password')
            
            if not request.user.check_password(current_password):
                messages.error(request, 'Текущий пароль введен неверно')
                return redirect('accounts:client_settings')
            
            if not new_password:
                messages.error(request, 'Введите новый пароль')
                return redirect('accounts:client_settings')
            
            if new_password != confirm_password:
                messages.error(request, 'Новый пароль и подтверждение не совпадают')
                return redirect('accounts:client_settings')
            
            if len(new_password) < 8:
                messages.error(request, 'Пароль должен содержать минимум 8 символов')
                return redirect('accounts:client_settings')
            
            request.user.set_password(new_password)
            request.user.save()
            
            messages.success(request, 'Пароль успешно изменен. Пожалуйста, войдите снова.')
            return redirect('accounts:login')
    
    context = {
        'user': request.user,
        'client': client,
        'active_tab': 'settings',
    }
    return render(request, 'accounts/client_settings.html', context)

@login_required
def ai_assistant(request):
    if request.method == "POST":
        try:
            data = json.loads(request.body)
            message = data.get("message", "").lower()
            print(f"Получен запрос: {message}")
            
             # Сохраняем вопрос в бд
            try:
                AIChatLog.objects.create(
                    user=request.user,
                    question=message
                )
                print(f"Вопрос сохранен в БД")
            except Exception as e:
                print(f"Ошибка сохранения вопроса: {e}")

            # Ищем номер заказа
            match = re.search(r'\d+', message)

            if "заказ" in message and match:
                order_id = int(match.group())
                print(f"Ищем заказ #{order_id}")

                try:
                    order = Order.objects.get(
                        id=order_id,
                        client=request.user.client_profile
                    )
                    print(f"Заказ найден, статус: {order.status}")

                    # Перевод статуса на русский
                    status_ru = {
                        'created': 'создан',
                        'pending': 'ожидает подтверждения курьером',
                        'assigned': 'назначен курьер',
                        'in_progress': 'в пути',
                        'delivered': 'доставлен',
                        'cancelled': 'отменён'
                    }.get(order.status, order.status)

                    prompt = f"""
Ты AI ассистент службы доставки АВН Бизнес Курьер.

Пользователь спрашивает о заказе №{order.id}.

Информация о заказе:
- Номер заказа: {order.id}
- Статус: {status_ru}
- Адрес отправки: {order.pickup_address}
- Адрес доставки: {order.delivery_address}
- Тип заказа: {order.get_order_type_display()}
- Вес: {order.weight if order.weight else 'не указан'} кг

Ответь пользователю вежливо, кратко и информативно. Сообщи статус заказа и основную информацию.
"""
                except Order.DoesNotExist:
                    print(f"Заказ №{order_id} не найден")
                    return JsonResponse({
                        "answer": f"Заказ №{order_id} не найден. Проверьте номер заказа или обратитесь в поддержку."
                    })
            else:
                # Обычные вопросы
                prompt = f"""
Ты AI ассистент службы доставки АВН Бизнес Курьер.

Твоя задача - помогать клиентам с вопросами о:
- Статусе заказов
- Тарифах доставки
- Стоимости услуг
- Сроках доставки
- Правилах оформления заказа

Отвечай вежливо, кратко (2-3 предложения), по делу.
Если не знаешь ответа, предложи обратиться в поддержку по телефону +7 (980) 888-88-88 или email abn-business@mail.ru

Вопрос клиента: {message}
"""

            print(f"Отправляем запрос в Ollama")
            
            # Запрос к Ollama
            response = requests.post(
                "http://localhost:11434/api/generate",
                json={
                    "model": "qwen2.5:3b",
                    "prompt": prompt,
                    "stream": False,
                    "options": {
                        "temperature": 0.7,
                        "max_tokens": 500
                    }
                },
                timeout=30
            )
            
            print(f"Статус ответа Ollama: {response.status_code}")
            
            if response.status_code == 200:
                answer = response.json().get("response", "Извините, не удалось получить ответ.")
                print(f"Ответ AI получен, длина: {len(answer)} символов")
                return JsonResponse({"answer": answer})
            else:
                return JsonResponse({
                    "answer": f"Извините, AI сервер временно недоступен. Пожалуйста, попробуйте позже."
                })
                
        except requests.exceptions.ConnectionError:
            print("Не удалось подключиться к Ollama")
            return JsonResponse({
                "answer": "AI сервер не запущен. Пожалуйста, сообщите об этом администратору."
            })
        except requests.exceptions.Timeout:
            print("Таймаут подключения к Ollama")
            return JsonResponse({
                "answer": "AI сервер не отвечает. Пожалуйста, попробуйте позже."
            })
        except Exception as e:
            print(f"Общая ошибка: {str(e)}")
            import traceback
            traceback.print_exc()
            return JsonResponse({"answer": f"Произошла ошибка. Пожалуйста, попробуйте позже."}, status=500)

    return render(request, 'accounts/ai_assistant.html')

def privacy_policy(request):
    "Страница с политикой конфиденциальности"
    return render(request, 'accounts/privacy_policy.html')

# ЛК Менеджер
@login_required
def manager_dashboard(request):
    """Главная страница менеджера"""
    if request.user.role != 'manager':
        messages.error(request, 'У вас нет доступа к этой странице')
        return redirect('accounts:home')
    
    # Активные заказы
    active_orders = Order.objects.filter(
        status__in=['created', 'pending', 'assigned', 'in_progress']
    ).count()
    
    # Новые заказы (ожидают назначения)
    new_orders = Order.objects.filter(status='pending').count()
    
    # Просроченные заказы (желаемая дата доставки уже прошла)
    from django.utils import timezone
    delayed_orders = Order.objects.filter(
        requested_delivery_date__lt=timezone.now().date(),
        status__in=['created', 'pending', 'assigned', 'in_progress']
    ).count()
    
    # Данные для графика за неделю
    today = timezone.now().date()
    week_start = today - timedelta(days=today.weekday())
    weekly_data = []
    for i in range(7):
        day = week_start + timedelta(days=i)
        count = Order.objects.filter(created_at__date=day).count()
        weekly_data.append(count)
    
    # Топ курьеров по рейтингу
    top_couriers = []
    for courier in Courier.objects.filter(user__role='courier').order_by('-avg_rating')[:3]:
        deliveries_count = Order.objects.filter(courier=courier, status='delivered').count()
        top_couriers.append({
            'name': courier.user.get_full_name(),
            'deliveries': deliveries_count,
            'avg_time': 30,
            'rating': float(courier.avg_rating)
        })
    
    context = {
        'active_orders': active_orders,
        'new_orders': new_orders,
        'delayed_orders': delayed_orders,
        'weekly_data': weekly_data,
        'top_couriers': top_couriers,
    }
    return render(request, 'accounts/manager_dashboard.html', context)

@login_required
def manager_notifications(request):
    """API для получения уведомлений менеджера"""
    if request.user.role != 'manager':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    notifications = []
    
    # 1. Заказы без курьера (created или pending)
    orders_without_courier = Order.objects.filter(
        status__in=['created', 'pending']
    )
    
    for order in orders_without_courier[:5]:
        notifications.append({
            'type': 'warning',
            'message': f'Заказ №{order.id} ({order.get_status_display()}) ожидает назначения курьера',
            'time': order.created_at.strftime('%H:%M'),
            'order_id': order.id
        })
    
    # 2. Просроченные заказы
    from django.utils import timezone
    delayed_orders = Order.objects.filter(
        requested_delivery_date__lt=timezone.now().date(),
        status__in=['created', 'pending', 'assigned', 'in_progress']
    )
    
    for order in delayed_orders[:5]:
        days_delayed = (timezone.now().date() - order.requested_delivery_date).days
        notifications.append({
            'type': 'danger',
            'message': f'⚠️ Заказ №{order.id} просрочен на {days_delayed} дн.',
            'time': order.created_at.strftime('%H:%M'),
            'order_id': order.id
        })
    
    return JsonResponse({'notifications': notifications})

@login_required
def manager_tasks(request):
    if request.user.role != 'manager':
        messages.error(request, 'У вас нет доступа к этой странице')
        return redirect('accounts:home')
    
    orders = Order.objects.exclude(
        status__in=['delivered', 'cancelled']
    ).order_by('-created_at')
    
    for order in orders:
        last_status = OrderStatusHistory.objects.filter(order=order).order_by('-changed_at').first()
        order.last_status_text = last_status.get_status_display() if last_status else order.get_status_display()
        order.last_status_time = last_status.changed_at if last_status else order.created_at

    # Получаем всех курьеров на смене с их рабочим временем
    available_couriers = Courier.objects.filter(shift_status='on').select_related('user')
    
    context = {
        'orders': orders,
        'available_couriers': available_couriers,
    }
    return render(request, 'accounts/manager_tasks.html', context)

@login_required
def manager_orders(request):
    "Страница завершенных заказов"
    if request.user.role != 'manager':
        messages.error(request, 'У вас нет доступа к этой странице')
        return redirect('accounts:home')
    
    # Базовый запрос - только завершенные заказы (доставленные)
    orders = Order.objects.filter(status='delivered').order_by('-delivered_at', '-created_at')
    
    # Поиск по названию компании клиента
    search_query = request.GET.get('search', '')
    if search_query:
        orders = orders.filter(
            Q(client__company_name__icontains=search_query) |
            Q(client__contact_person_last_name__icontains=search_query) |
            Q(recipient_last_name__icontains=search_query) |
            Q(recipient_company__icontains=search_query)
        )
    
    # Фильтр по дате доставки
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    
    # Фильтр дат отдельно от поиска
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            orders = orders.filter(delivered_at__date__gte=date_from_obj)
        except (ValueError, TypeError):
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            orders = orders.filter(delivered_at__date__lte=date_to_obj)
        except (ValueError, TypeError):
            pass
    
    # Пагинация
    paginator = Paginator(orders, 4)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'date_from': date_from,
        'date_to': date_to,
        'total_count': orders.count(),
    }
    return render(request, 'accounts/manager_orders.html', context)


@login_required
def manager_couriers(request):
    "Страница управления курьерами"
    if request.user.role != 'manager':
        messages.error(request, 'У вас нет доступа к этой странице')
        return redirect('accounts:home')
    
    couriers = Courier.objects.select_related('user').all().order_by('-user__date_joined')
    
    # Поиск по ФИО
    search_query = request.GET.get('search', '')
    if search_query:
        couriers = couriers.filter(
            Q(user__first_name__icontains=search_query) |
            Q(user__last_name__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )
    
    # Пагинация
    paginator = Paginator(couriers, 4)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'total_count': couriers.count(),
    }
    return render(request, 'accounts/manager_couriers.html', context)

@login_required
def manager_courier_add(request):
    "Добавление нового курьера менеджером"
    if request.user.role != 'manager':
        messages.error(request, 'У вас нет доступа к этой странице')
        return redirect('accounts:home')
    
    if request.method == 'POST':
        # Генерируем временный пароль
        temp_password = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
        
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        
        # Создаем пользователя
        user = User.objects.create(
            email=email,
            first_name=first_name,
            last_name=last_name,
            patronymic=request.POST.get('patronymic', ''),
            phone=request.POST.get('phone'),
            role='courier',
            password=make_password(temp_password)
        )
        
        # Создаем профиль курьера
        from django.utils import timezone
        Courier.objects.create(
            user=user,
            hire_date=request.POST.get('hire_date') or timezone.now().date(),
            position=request.POST.get('position', 'Курьер'),
            shift_status=request.POST.get('shift_status', 'off'),
            citizenship=request.POST.get('citizenship'),
            passport_series=request.POST.get('passport_series'),
            passport_number=request.POST.get('passport_number'),
            passport_department_code=request.POST.get('passport_department_code'),
            passport_issued_by=request.POST.get('passport_issued_by'),
            passport_issue_date=request.POST.get('passport_issue_date'),
            registration_address=request.POST.get('registration_address'),
            actual_address=request.POST.get('actual_address'),
        )
        
        # Отправляем email с паролем
        try:
            send_mail(
                subject='Добро пожаловать в АВН Бизнес Курьер!',
                message=f"""
            Здравствуйте, {first_name} {last_name}!

            Вы были добавлены в систему АВН Бизнес Курьер в качестве курьера.

            Ваши данные для входа:
            Email: {email}
            Временный пароль: {temp_password}

            Пожалуйста, войдите в систему и смените пароль при первом входе.

            Ссылка для входа: http://127.0.0.1:8000/login/

            С уважением,
            Команда АВН Бизнес Курьер
                """,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=False,
            )
            messages.success(request, f'Курьер {email} успешно создан. Пароль отправлен на почту.')
        except Exception as e:
            messages.warning(request, f'Курьер создан, но не удалось отправить email: {e}. Временный пароль: {temp_password}')
        
        return redirect('accounts:manager_couriers')
    
    context = {
        'title': 'Добавить курьера',
        'courier_user': None,
        'courier': None,
    }
    return render(request, 'accounts/manager_courier_form.html', context)

@login_required
def manager_courier_edit(request, user_id):
    "Редактирование курьера"
    if request.user.role != 'manager':
        messages.error(request, 'У вас нет доступа к этой странице')
        return redirect('accounts:home')
    
    try:
        user = User.objects.get(id=user_id, role='courier')
        courier = user.courier_profile
    except (User.DoesNotExist, Courier.DoesNotExist):
        messages.error(request, 'Курьер не найден')
        return redirect('accounts:manager_couriers')
    
    if request.method == 'POST':
        # Обновляем пользователя
        user.first_name = request.POST.get('first_name')
        user.last_name = request.POST.get('last_name')
        user.patronymic = request.POST.get('patronymic', '')
        user.phone = request.POST.get('phone')
        user.save()
        
        # Обновляем курьера
        courier.hire_date = request.POST.get('hire_date')
        courier.position = request.POST.get('position')
        courier.shift_status = request.POST.get('shift_status')
        courier.citizenship = request.POST.get('citizenship')
        courier.passport_series = request.POST.get('passport_series')
        courier.passport_number = request.POST.get('passport_number')
        courier.passport_department_code = request.POST.get('passport_department_code')
        courier.passport_issued_by = request.POST.get('passport_issued_by')
        courier.passport_issue_date = request.POST.get('passport_issue_date')
        courier.registration_address = request.POST.get('registration_address')
        courier.actual_address = request.POST.get('actual_address')
        courier.save()
        
        messages.success(request, 'Данные курьера обновлены')
        return redirect('accounts:manager_couriers')
    
    context = {
        'title': 'Редактировать курьера',
        'courier_user': user, 
        'courier': courier,
    }
    return render(request, 'accounts/manager_courier_form.html', context)

@login_required
def manager_courier_detail(request, user_id):
    "Получение деталей курьера - модальное окно"
    if request.user.role != 'manager':
        return JsonResponse({'error': 'Доступ запрещен'}, status=403)
    
    try:
        user = User.objects.get(id=user_id, role='courier')
        courier = user.courier_profile
        
        data = {
            'id': user.id,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'patronymic': user.patronymic or '',
            'phone': user.phone or '',
            'date_joined': user.date_joined.strftime('%d.%m.%Y %H:%M'),
            'hire_date': courier.hire_date.strftime('%d.%m.%Y'),
            'position': courier.position,
            'shift_status': courier.get_shift_status_display(),
            'citizenship': courier.citizenship,
            'passport_series': courier.passport_series,
            'passport_number': courier.passport_number,
            'passport_department_code': courier.passport_department_code,
            'passport_issued_by': courier.passport_issued_by,
            'passport_issue_date': courier.passport_issue_date.strftime('%d.%m.%Y'),
            'registration_address': courier.registration_address,
            'actual_address': courier.actual_address,
            'total_orders': courier.total_orders,
            'avg_rating': float(courier.avg_rating),
            'total_work_hours': float(courier.total_work_hours),
        }
        return JsonResponse({'success': True, 'data': data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=404)

@login_required
def manager_clients(request):
    "Страница управления клиентами"
    if request.user.role != 'manager':
        messages.error(request, 'У вас нет доступа к этой странице')
        return redirect('accounts:home')
    
    clients = Client.objects.select_related('user').annotate(
        orders_count=Count('orders')).order_by('-user__date_joined')
    
    # Поиск
    search_query = request.GET.get('search', '')
    if search_query:
        clients = clients.filter(
            Q(company_name__icontains=search_query) |
            Q(inn__icontains=search_query) |
            Q(contact_person_last_name__icontains=search_query) |
            Q(contact_person_first_name__icontains=search_query) |
            Q(user__email__icontains=search_query)
        )
    
    # Пагинация
    paginator = Paginator(clients, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Дополнительная статистика
    from django.utils import timezone
    total_orders_count = Order.objects.filter(status='delivered').count()
    new_clients_this_month = Client.objects.filter(
        user__date_joined__month=timezone.now().month
    ).count()
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'total_count': clients.count(),
        'total_orders_count': total_orders_count,
        'new_clients_this_month': new_clients_this_month,
    }
    return render(request, 'accounts/manager_clients.html', context)

@login_required
def manager_client_add(request):
    "Добавление нового клиента менеджером"
    if request.user.role != 'manager':
        messages.error(request, 'У вас нет доступа к этой странице')
        return redirect('accounts:home')
    
    if request.method == 'POST':
        # Генеррация временного пароля
        temp_password = ''.join(random.choices(string.ascii_letters + string.digits, k=10))
        email = request.POST.get('email')
        first_name = request.POST.get('first_name')
        last_name = request.POST.get('last_name')
        
        # Создаем пользователя
        user = User.objects.create(
            email=email,
            first_name=first_name,
            last_name=last_name,
            patronymic=request.POST.get('patronymic', ''),
            phone=request.POST.get('phone'),
            role='client',
            password=make_password(temp_password)
        )
        
        # Создаем профиль клиента
        Client.objects.create(
            user=user,
            company_name=request.POST.get('company_name'),
            inn=request.POST.get('inn'),
            kpp=request.POST.get('kpp'),
            legal_address=request.POST.get('legal_address'),
            actual_address=request.POST.get('actual_address'),
            company_phone=request.POST.get('company_phone'),
            company_email=request.POST.get('company_email'),
            bank=request.POST.get('bank'),
            settlement_account=request.POST.get('settlement_account'),
            correspondent_account=request.POST.get('correspondent_account'),
            contact_person_first_name=request.POST.get('contact_person_first_name'),
            contact_person_last_name=request.POST.get('contact_person_last_name'),
            contact_person_patronymic=request.POST.get('contact_person_patronymic', ''),
            contact_person_phone=request.POST.get('contact_person_phone'),
            contact_person_email=request.POST.get('contact_person_email'),
        )
        
        # Отправка письма (email) с паролем
        try:
            send_mail(
                subject='Добро пожаловать в АВН Бизнес Курьер!',
                message=f"""
                Здравствуйте, {first_name} {last_name}!

                Для вас был создан аккаунт в системе АВН Бизнес Курьер.

                Ваши данные для входа:
                Email: {email}
                Временный пароль: {temp_password}

                Пожалуйста, войдите в систему и смените пароль при первом входе.

                Ссылка для входа: http://127.0.0.1:8000/login/

                С уважением,
                Команда АВН Бизнес Курьер
                """,
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=False,
            )
            messages.success(request, f'Клиент {email} успешно создан. Пароль отправлен на почту.')
        except Exception as e:
            messages.warning(request, f'Клиент создан, но не удалось отправить email: {e}. Временный пароль: {temp_password}')
        
        return redirect('accounts:manager_clients')
    
    context = {
        'title': 'Добавить клиента',
        'client_user': None,
        'client': None,
    }
    return render(request, 'accounts/manager_client_form.html', context)


@login_required
def manager_client_edit(request, user_id):
    "Редактирование клиента"
    if request.user.role != 'manager':
        messages.error(request, 'У вас нет доступа к этой странице')
        return redirect('accounts:home')
    
    try:
        user = User.objects.get(id=user_id, role='client')
        client = user.client_profile
    except (User.DoesNotExist, Client.DoesNotExist):
        messages.error(request, 'Клиент не найден')
        return redirect('accounts:manager_clients')
    
    if request.method == 'POST':
        # Обновлени пользователя
        user.first_name = request.POST.get('first_name')
        user.last_name = request.POST.get('last_name')
        user.patronymic = request.POST.get('patronymic', '')
        user.phone = request.POST.get('phone')
        user.save()
        
        # Обновление клиента
        client.company_name = request.POST.get('company_name')
        client.inn = request.POST.get('inn')
        client.kpp = request.POST.get('kpp')
        client.legal_address = request.POST.get('legal_address')
        client.actual_address = request.POST.get('actual_address')
        client.company_phone = request.POST.get('company_phone')
        client.company_email = request.POST.get('company_email')
        client.bank = request.POST.get('bank')
        client.settlement_account = request.POST.get('settlement_account')
        client.correspondent_account = request.POST.get('correspondent_account')
        client.contact_person_first_name = request.POST.get('contact_person_first_name')
        client.contact_person_last_name = request.POST.get('contact_person_last_name')
        client.contact_person_patronymic = request.POST.get('contact_person_patronymic', '')
        client.contact_person_phone = request.POST.get('contact_person_phone')
        client.contact_person_email = request.POST.get('contact_person_email')
        client.save()
        
        messages.success(request, 'Данные клиента обновлены')
        return redirect('accounts:manager_clients')
    
    context = {
        'title': 'Редактировать клиента',
        'client_user': user,
        'client': client,
    }
    return render(request, 'accounts/manager_client_form.html', context)


@login_required
def manager_client_detail(request, user_id):
    "Получение деталей клиента для модального окна"
    if request.user.role != 'manager':
        return JsonResponse({'error': 'Доступ запрещен'}, status=403)
    
    try:
        user = User.objects.get(id=user_id, role='client')
        client = user.client_profile
        orders_count = Order.objects.filter(client=client).count()
        
        data = {
            'id': user.id,
            'email': user.email,
            'first_name': user.first_name,
            'last_name': user.last_name,
            'patronymic': user.patronymic or '',
            'phone': user.phone or '',
            'date_joined': user.date_joined.strftime('%d.%m.%Y %H:%M'),
            'company_name': client.company_name,
            'inn': client.inn,
            'kpp': client.kpp,
            'legal_address': client.legal_address,
            'actual_address': client.actual_address,
            'company_phone': client.company_phone,
            'company_email': client.company_email,
            'bank': client.bank,
            'settlement_account': client.settlement_account,
            'correspondent_account': client.correspondent_account,
            'contact_person_first_name': client.contact_person_first_name,
            'contact_person_last_name': client.contact_person_last_name,
            'contact_person_patronymic': client.contact_person_patronymic or '',
            'contact_person_phone': client.contact_person_phone,
            'contact_person_email': client.contact_person_email,
            'orders_count': orders_count,
        }
        return JsonResponse({'success': True, 'data': data})
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=404)

@login_required
def manager_reports(request):
    "Страница отчетов менеджера"
    if request.user.role != 'manager':
        messages.error(request, 'У вас нет доступа к этой странице')
        return redirect('accounts:home')
    
    # Получаем параметры фильтрации
    report_type = request.GET.get('report_type', 'orders')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    format_type = request.GET.get('format', '')

    # Если запрошено скачивание
    if format_type:
        return generate_report(request, report_type, date_from, date_to, format_type)
    
    # Статистика для отображения на странице
    today = timezone.now().date()
    start_of_month = today.replace(day=1)
    end_of_month = (start_of_month + timedelta(days=32)).replace(day=1) - timedelta(days=1)
    
    # Общая статистика
    total_orders = Order.objects.count()
    delivered_orders = Order.objects.filter(status='delivered').count()
    active_orders = Order.objects.filter(status__in=['created', 'pending', 'assigned', 'in_progress']).count()
    cancelled_orders = Order.objects.filter(status='cancelled').count()
    
    # Статистика за месяц
    monthly_orders = Order.objects.filter(created_at__date__gte=start_of_month, created_at__date__lte=end_of_month).count()
    monthly_delivered = Order.objects.filter(status='delivered', delivered_at__date__gte=start_of_month, delivered_at__date__lte=end_of_month).count()
    
    # Статистика по типам заказов
    orders_by_type = Order.objects.values('order_type').annotate(count=Count('id'))
    
    # Статистика по курьерам (топ-5)
    top_couriers = Courier.objects.annotate(
        deliveries=Count('orders', filter=Q(orders__status='delivered'))
    ).order_by('-deliveries')[:5]
    
    # Статистика по дням для графика
    daily_stats = []
    for i in range(30):
        day = today - timedelta(days=29-i)
        count = Order.objects.filter(created_at__date=day).count()
        daily_stats.append({
            'date': day.strftime('%d.%m'),
            'count': count
        })
    
    context = {
        'active_tab': 'reports',
        'report_type': report_type,
        'date_from': date_from,
        'date_to': date_to,
        'total_orders': total_orders,
        'delivered_orders': delivered_orders,
        'active_orders': active_orders,
        'cancelled_orders': cancelled_orders,
        'monthly_orders': monthly_orders,
        'monthly_delivered': monthly_delivered,
        'orders_by_type': orders_by_type,
        'top_couriers': top_couriers,
        'daily_stats': daily_stats,
        'start_of_month': start_of_month.strftime('%d.%m.%Y'),
        'end_of_month': end_of_month.strftime('%d.%m.%Y'),
    }
    return render(request, 'accounts/manager_reports.html', context)

def generate_report(request, report_type, date_from, date_to, format_type):
    "Генерация отчета в Excel или PDF"
    
    # Фильтрация по датам
    orders = Order.objects.select_related('client', 'courier__user').all().order_by('-created_at')
    
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__gte=date_from_obj)
        except:
            pass
    
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d').date()
            orders = orders.filter(created_at__date__lte=date_to_obj)
        except:
            pass
    
    # Формируем данные для отчета
    data = []
    for order in orders:
        data.append({
            'id': order.id,
            'created_at': order.created_at.strftime('%d.%m.%Y %H:%M'),
            'client_name': order.client.company_name,
            'order_type': order.get_order_type_display(),
            'tariff': order.get_tariff_display(),
            'pickup_address': order.pickup_address[:100],
            'delivery_address': order.delivery_address[:100],
            'weight': str(order.weight) if order.weight else '—',
            'status': order.get_status_display(),
            'courier_name': order.courier.user.get_full_name() if order.courier else 'Не назначен',
            'delivered_at': order.delivered_at.strftime('%d.%m.%Y %H:%M') if order.delivered_at else '—',
        })
    
    if format_type == 'excel':
        return generate_excel_report(data, report_type, date_from, date_to)
    elif format_type == 'pdf':
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="report.pdf"'

        font_path = os.path.join(settings.BASE_DIR, 'static', 'fonts', 'DejaVuSans.ttf')

        pdfmetrics.registerFont(TTFont('DejaVuSans', font_path))

        doc = SimpleDocTemplate(
            response,
            pagesize=landscape(A4),
            leftMargin=20,
            rightMargin=20,
            topMargin=20,
            bottomMargin=20
        )

        elements = []
        styles = getSampleStyleSheet()

        # Цвет компании
        company_color = colors.HexColor("#1E7A57")

        # Стили
        title_style = ParagraphStyle(
            'title',
            parent=styles['Heading1'],
            fontName='DejaVuSans',
            fontSize=20,
            leading=24,
            textColor=company_color,
            alignment=1,
            spaceAfter=10
        )

        subtitle_style = ParagraphStyle(
            'subtitle',
            parent=styles['Normal'],
            fontName='DejaVuSans',
            fontSize=10,
            alignment=1,
            textColor=colors.grey,
            spaceAfter=20
        )

        normal_style = ParagraphStyle(
            'normal',
            parent=styles['Normal'],
            fontName='DejaVuSans',
            fontSize=9,
            leading=11
        )

        footer_style = ParagraphStyle(
            'footer',
            parent=styles['Normal'],
            fontName='DejaVuSans',
            fontSize=8,
            alignment=1,
            textColor=colors.grey
        )

        # Заголовок
        elements.append(Paragraph("АВН Бизнес Курьер", title_style))
        elements.append(Paragraph("Отчет по заказам и аналитике", subtitle_style))

        # Статистика
        total_orders = Order.objects.count()
        delivered = Order.objects.filter(status='delivered').count()
        active = Order.objects.exclude(status__in=['delivered', 'cancelled']).count()
        cancelled = Order.objects.filter(status='cancelled').count()

        stats_data = [
            ['Показатель', 'Значение'],
            ['Всего заказов', total_orders],
            ['Доставлено', delivered],
            ['Активных', active],
            ['Отменено', cancelled],
        ]

        stats_table = Table(stats_data, colWidths=[250, 150])
        stats_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), company_color),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,-1), 'DejaVuSans'),
            ('GRID', (0,0), (-1,-1), 1, colors.lightgrey),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#F5F3F0')]),
            ('ALIGN', (1,1), (1,-1), 'CENTER'),
        ]))

        elements.append(stats_table)
        elements.append(Spacer(1, 20))

        # Основная таблица
        table_data = [[
            '№',
            'Дата',
            'Клиент',
            'Тип',
            'Тариф',
            'Адрес доставки',
            'Вес',
            'Статус',
            'Курьер'
        ]]

        for item in data:
            table_data.append([
                item['id'],
                item['created_at'],
                item['client_name'],
                item['order_type'],
                item['tariff'],
                item['delivery_address'][:35],
                item['weight'],
                item['status'],
                item['courier_name']
            ])

        main_table = Table(
            table_data,
            colWidths=[35, 75, 110, 70, 70, 180, 45, 75, 110]
        )

        main_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), company_color),
            ('TEXTCOLOR', (0,0), (-1,0), colors.white),
            ('FONTNAME', (0,0), (-1,-1), 'DejaVuSans'),
            ('FONTSIZE', (0,0), (-1,-1), 8),
            ('GRID', (0,0), (-1,-1), 0.5, colors.lightgrey),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [
                colors.white,
                colors.HexColor('#F8F9FA')
            ]),
        ]))

        elements.append(main_table)
        elements.append(Spacer(1, 25))

        # Подвал
        elements.append(Paragraph(
            "АВН Бизнес Курьер | avn@delivery.ru | +7 (999) 123-45-67",
            footer_style
        ))

        doc.build(elements)

        return response

def generate_excel_report(data, report_type, date_from, date_to):
        "Генерация Excel отчета"
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Отчет_{report_type}"
        
        # Стили
        header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='1E7A57', end_color='1E7A57', fill_type='solid')
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        # Заголовки столбцов
        headers = [
            '№ заказа', 'Дата создания', 'Клиент', 'Тип заказа', 'Тариф',
            'Адрес отправки', 'Адрес доставки', 'Вес (кг)', 'Статус', 'Курьер', 'Дата доставки'
        ]
        
        # Заполняем заголовки
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Заполняем данные
        for row, item in enumerate(data, 2):
            ws.cell(row=row, column=1, value=item['id'])
            ws.cell(row=row, column=2, value=item['created_at'])
            ws.cell(row=row, column=3, value=item['client_name'])
            ws.cell(row=row, column=4, value=item['order_type'])
            ws.cell(row=row, column=5, value=item['tariff'])
            ws.cell(row=row, column=6, value=item['pickup_address'])
            ws.cell(row=row, column=7, value=item['delivery_address'])
            ws.cell(row=row, column=8, value=item['weight'])
            ws.cell(row=row, column=9, value=item['status'])
            ws.cell(row=row, column=10, value=item['courier_name'])
            ws.cell(row=row, column=11, value=item['delivered_at'])
        
        # Автоширина колонок
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            ws.column_dimensions[column_letter].width = 20
        
        # Формируем имя файла
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"Отчет_{report_type}_{timestamp}.xlsx"
        
        # Создаем HTTP ответ
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        wb.save(response)
        return response

@login_required
def manager_settings(request):
        "Настройки менеджера - изменение профиля и пароля"
        if request.user.role != 'manager':
            messages.error(request, 'У вас нет доступа к этой странице')
            return redirect('accounts:home')
        
        if request.method == 'POST':
            action = request.POST.get('action')
            
            # Обновление профиля
            if action == 'update_profile':
                user = request.user
                user.first_name = request.POST.get('first_name', '').strip()
                user.last_name = request.POST.get('last_name', '').strip()
                user.patronymic = request.POST.get('patronymic', '').strip()
                user.phone = request.POST.get('phone', '').strip()
                user.save()
                
                messages.success(request, 'Данные профиля успешно обновлены')
                return redirect('accounts:manager_settings')
            
            # Смена пароля
            elif action == 'change_password':
                current_password = request.POST.get('current_password')
                new_password = request.POST.get('new_password')
                confirm_password = request.POST.get('confirm_password')
                
                # Проверяем текущий пароль
                if not request.user.check_password(current_password):
                    messages.error(request, 'Текущий пароль введен неверно')
                    return redirect('accounts:manager_settings')
                
                # Проверяем, что новый пароль не пустой
                if not new_password:
                    messages.error(request, 'Введите новый пароль')
                    return redirect('accounts:manager_settings')
                
                # Проверяем совпадение паролей
                if new_password != confirm_password:
                    messages.error(request, 'Новый пароль и подтверждение не совпадают')
                    return redirect('accounts:manager_settings')
                
                # Проверяем длину пароля
                if len(new_password) < 8:
                    messages.error(request, 'Пароль должен содержать минимум 8 символов')
                    return redirect('accounts:manager_settings')
                
                # Меняем пароль
                request.user.set_password(new_password)
                request.user.save()
                
                messages.success(request, 'Пароль успешно изменен. Пожалуйста, войдите снова.')
                return redirect('accounts:login')
        
        context = {
            'user': request.user,
        }
        return render(request, 'accounts/manager_settings.html', context)

@login_required
def manager_ai_stats(request):
        "Страница статистики AI чата для менеджера"
        if request.user.role != 'manager':
            messages.error(request, 'У вас нет доступа к этой странице')
            return redirect('accounts:home')
        
        context = {
            'active_tab': 'ai_stats',
        }
        return render(request, 'accounts/manager_ai_stats.html', context)

@login_required
def manager_tasks_count(request):
        "API для получения количества задач менеджера"
        if request.user.role != 'manager':
            return JsonResponse({'error': 'Access denied'}, status=403)
        
        # Заказы без курьера (created или pending)
        orders_without_courier = Order.objects.filter(
            status__in=['created', 'pending']
        ).count()
        
        # Просроченные заказы
        from django.utils import timezone
        delayed_orders = Order.objects.filter(
            requested_delivery_date__lt=timezone.now().date(),
            status__in=['created', 'pending', 'assigned', 'in_progress']
        ).count()
        
        # Общее количество задач
        total_tasks = orders_without_courier + delayed_orders
        
        return JsonResponse({'count': total_tasks})

@login_required
def assign_courier_ajax(request):
    if request.user.role != 'manager':
        return JsonResponse({'error': 'Доступ запрещен'}, status=403)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            order_id = data.get('order_id')
            user_id = data.get('courier_id')
            
            order = get_object_or_404(Order, pk=order_id)
            courier = get_object_or_404(Courier, user_id=user_id)
            
            # Проверки
            if order.status not in ['created', 'pending']:
                return JsonResponse({'error': 'Нельзя назначить курьера'}, status=400)
            
            if not courier.is_available():
                return JsonResponse({'error': 'Курьер недоступен'}, status=400)
            
            # Назначение
            order.courier = courier
            order.status = 'pending'
            order.save()
            
            # История
            OrderStatusHistory.objects.create(
                order=order,
                status='pending',
                comment=f'Назначен курьер {courier.user.get_full_name()}, ожидает подтверждения'
            )
            
            # Уведомление
            CourierNotification.objects.create(
                courier=courier,
                order=order,
                message=f'Вам назначен заказ №{order.id}',
                notification_type='new'
            )
            
            return JsonResponse({'success': True})
            
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Метод не разрешен'}, status=405)

@login_required
def delete_order_ajax(request):
    "AJAX удаление заказа"
    if request.user.role != 'manager':
        return JsonResponse({'error': 'Доступ запрещен'}, status=403)
    
    if request.method == 'POST':
        data = json.loads(request.body)
        order_id = data.get('order_id')
        
        try:
            order = Order.objects.get(id=order_id)
            order.delete()
            return JsonResponse({'success': True, 'message': 'Заказ удален'})
        except Order.DoesNotExist:
            return JsonResponse({'error': 'Заказ не найден'}, status=404)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Метод не разрешен'}, status=405)

@login_required
def get_order_details(request):
    "Получение деталей заказа для модального окна"
    # Разрешаем доступ и менеджеру, и клиенту (владельцу заказа)
    order_id = request.GET.get('order_id')
    
    try:
        order = Order.objects.get(id=order_id)
        
        # Проверка прав доступа:
        # - Менеджер может смотреть любой заказ
        # - Клиент может смотреть только свои заказы
        if request.user.role == 'manager':
            pass  # менеджеру всё можно
        elif request.user.role == 'client':
            # Проверяем, что заказ принадлежит этому клиенту
            if order.client != request.user.client_profile:
                return JsonResponse({'error': 'Доступ запрещен'}, status=403)
        else:
            return JsonResponse({'error': 'Доступ запрещен'}, status=403)
        
        # Получаем всю историю статусов в хронологическом порядке
        status_history = OrderStatusHistory.objects.filter(order=order).order_by('changed_at')
        history_list = []
        for status in status_history:
            history_list.append({
                'status': status.get_status_display(),
                'time': status.changed_at.strftime('%d.%m.%Y %H:%M'),
                'comment': status.comment
            })
        
        # Получаем последний статус
        last_status = status_history.last()
        
        # Определяем статус оплаты из поля payment_status
        payment_status_display = dict(Order.PAYMENT_CHOICES).get(order.payment_status, 'Не указан')
        
        # Данные для клиента
        data = {
            'id': order.id,
            'client_name': order.client.company_name,
            'contact_person': order.client.get_contact_person_full_name(),
            'contact_phone': order.client.contact_person_phone,
            'pickup_address': order.pickup_address,
            'delivery_address': order.delivery_address,
            'order_type': order.get_order_type_display(),
            'tariff': order.get_tariff_display(),
            'weight': str(order.weight) if order.weight else 'не указан',
            'recipient_name': order.get_recipient_full_name(),
            'recipient_phone': order.recipient_phone,
            'recipient_company': order.recipient_company or '—',
            'created_at': order.created_at.strftime('%d.%m.%Y %H:%M'),
            'current_status': order.get_status_display(),
            'payment_status': payment_status_display, 
            'total_amount': str(order.total_amount) if order.total_amount else '0',
            'status_history': history_list,
            'client_comment': order.client_comment or 'Нет комментария',
        }
        
        # Добавляем информацию о курьере, если есть
        if order.courier:
            data['courier_name'] = order.courier.user.get_full_name()
            data['courier_phone'] = order.courier.user.phone or '—'
        else:
            data['courier_name'] = 'Не назначен'
            data['courier_phone'] = '—'
        
        return JsonResponse({'success': True, 'data': data})
        
    except Order.DoesNotExist:
        return JsonResponse({'error': 'Заказ не найден'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)
    
@login_required
def ai_chat_stats(request):
    "Статистика запросов к AI чату (менеджер)"
    if request.user.role != 'manager':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    from django.db.models import Count
    from datetime import timedelta
    
    # Популярные вопросы за последние 30 дней
    popular_questions = AIChatLog.objects.filter(
        created_at__gte=timezone.now() - timedelta(days=30)
    ).values('question').annotate(
        count=Count('id')
    ).order_by('-count')[:20]
    
    # Количество запросов по дням
    daily_stats = []
    for i in range(30):
        day = timezone.now().date() - timedelta(days=i)
        count = AIChatLog.objects.filter(created_at__date=day).count()
        daily_stats.append({'date': day.strftime('%d.%m'), 'count': count})
    
    return JsonResponse({
        'total_queries': AIChatLog.objects.count(),
        'popular_questions': list(popular_questions),
        'daily_stats': daily_stats
    })

@login_required
def ai_chat_stats(request):
    "Статистика запросов к AI чату (ЛК менеджер)"
    if request.user.role != 'manager':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    total_queries = AIChatLog.objects.count()
    queries_today = AIChatLog.objects.filter(created_at__date=timezone.now().date()).count()
    queries_week = AIChatLog.objects.filter(created_at__gte=timezone.now() - timedelta(days=7)).count()
    
    # Популярные вопросы за последние 30 дней
    popular_questions = AIChatLog.objects.filter(
        created_at__gte=timezone.now() - timedelta(days=30)
    ).values('question').annotate(
        count=Count('id')
    ).order_by('-count')[:10]
    
    # Статистика по дням для графика
    daily_stats = []
    for i in range(6, -1, -1):
        day = timezone.now().date() - timedelta(days=i)
        count = AIChatLog.objects.filter(created_at__date=day).count()
        daily_stats.append({
            'date': day.strftime('%d.%m'),
            'count': count
        })
    
    return JsonResponse({
        'total_queries': total_queries,
        'queries_today': queries_today,
        'queries_week': queries_week,
        'popular_questions': list(popular_questions),
        'daily_stats': daily_stats
    })

@login_required
def ai_chat_last_queries(request):
    "API для получения последних запросов к AI чату"
    if request.user.role != 'manager':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    from delivery.models import AIChatLog
    
    last_queries = AIChatLog.objects.select_related('user').order_by('-created_at')[:20]
    
    queries_data = []
    for log in last_queries:
        queries_data.append({
            'user_name': log.user.get_full_name() or log.user.email,
            'question': log.question[:100],
            'created_at': log.created_at.strftime('%d.%m.%Y %H:%M')
        })
    
    return JsonResponse({'queries': queries_data})

def main_page(request):
    "Главная страница сайта"
    return render(request, 'main_page.html')

@login_required
def client_orders(request):
    try:
        client = request.user.client_profile

        orders = Order.objects.filter(
            client=client,
            campaign_recipient__isnull=True
        ).order_by('-created_at')

        campaigns = Campaign.objects.filter(
            client=client
        ).prefetch_related('recipients').order_by('-created_at')

        # объединяем списки
        items = sorted(
            chain(orders, campaigns),
            key=attrgetter('created_at'),
            reverse=True
        )
        for obj in items:
            obj.is_campaign = isinstance(obj, Campaign)

        paginator = Paginator(items, 5)
        page_number = request.GET.get('page', 1)
        page_obj = paginator.get_page(page_number)

        active_count = orders.exclude(status='delivered').count()
        delivered_count = orders.filter(status='delivered').count()
        total_orders = orders.count() + campaigns.count()

    except:
        page_obj = []
        active_count = 0
        delivered_count = 0
        total_orders = 0

    return render(request, 'accounts/clients_orders.html', {
        'page_obj': page_obj,
        'active_count': active_count,
        'delivered_count': delivered_count,
        'total_orders': total_orders,
        'total_count': total_orders,
        'active_tab': 'orders',
    })


@login_required
def campaign_details_api(request):
    "API для получения деталей кампании"
    if request.user.role != 'client':
        return JsonResponse({'error': 'Access denied'}, status=403)
    
    campaign_id = request.GET.get('campaign_id')
    
    try:
        campaign = Campaign.objects.get(id=campaign_id, client=request.user.client_profile)
        
        data = {
            'id': campaign.id,
            'name': campaign.name,
            'occasion_display': dict(Campaign.OCCASION_CHOICES).get(campaign.occasion, '—'),
            'created_at': campaign.created_at.strftime('%d.%m.%Y %H:%M'),
            'pickup_address': campaign.pickup_address,
            'total_recipients': campaign.total_recipients,
            'delivered_count': campaign.recipients.filter(status='delivered').count(),
        }
        
        return JsonResponse({'success': True, 'data': data})
        
    except Campaign.DoesNotExist:
        return JsonResponse({'error': 'Кампания не найдена'}, status=404)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)